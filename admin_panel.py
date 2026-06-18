"""
LBAMS v4.3 Admin Panel
Run: python admin_panel.py
Default password: admin123
"""

from __future__ import annotations

import csv
import os
import shutil
from datetime import datetime
from pathlib import Path

import cv2

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    EXCEL_OK = True
except Exception:
    EXCEL_OK = False

import train_model

BASE_DIR = Path(__file__).resolve().parent
DATASET_DIR = BASE_DIR / "dataset"
ATTENDANCE_DIR = BASE_DIR / "attendance"
ADMIN_PASSWORD = "admin123"
PHOTOS_PER_ANGLE = 5
ANGLES = ["center", "left", "right", "up", "down"]

DATASET_DIR.mkdir(exist_ok=True)
ATTENDANCE_DIR.mkdir(exist_ok=True)


def clear():
    os.system("cls" if os.name == "nt" else "clear")


def pause():
    input("\nPress Enter to continue...")


def folder_name(name: str) -> str:
    cleaned = "_".join(name.strip().split())
    return "".join(ch for ch in cleaned if ch.isalnum() or ch in "_-" ).strip("_")


def display_name(folder: str) -> str:
    return folder.replace("_", " ")


def list_students():
    return [p for p in sorted(DATASET_DIR.iterdir()) if p.is_dir()]


def header(title="ADMIN PANEL"):
    clear()
    print("=" * 68)
    print(f" LBAMS v4.3 - {title}")
    print("=" * 68)


def capture_photos(student_folder: Path, student_name: str) -> bool:
    cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
    cap = cv2.VideoCapture(0, cv2.CAP_DSHOW if os.name == "nt" else 0)
    if not cap.isOpened():
        cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        print("Camera not found.")
        return False

    cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1280)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)
    instructions = {
        "center": "Look straight at the camera",
        "left": "Turn face slightly LEFT",
        "right": "Turn face slightly RIGHT",
        "up": "Look slightly UP",
        "down": "Look slightly DOWN",
    }
    saved_total = 0
    for angle in ANGLES:
        saved = 0
        print(f"\nAngle: {angle.upper()} - {instructions[angle]}")
        print("SPACE = capture | S = skip this angle | Q = quit")
        while saved < PHOTOS_PER_ANGLE:
            ok, frame = cap.read()
            if not ok:
                break
            show = frame.copy()
            small = cv2.resize(frame, (0, 0), fx=0.5, fy=0.5)
            gray = cv2.cvtColor(small, cv2.COLOR_BGR2GRAY)
            faces = cascade.detectMultiScale(gray, 1.1, 5, minSize=(60, 60))
            for (x, y, w, h) in faces:
                X, Y, W, H = int(x / 0.5), int(y / 0.5), int(w / 0.5), int(h / 0.5)
                cv2.rectangle(show, (X, Y), (X + W, Y + H), (0, 220, 0), 2)
            cv2.rectangle(show, (0, 0), (show.shape[1], 64), (20, 20, 20), cv2.FILLED)
            cv2.putText(show, f"{student_name} | {angle.upper()} | {saved}/{PHOTOS_PER_ANGLE}", (12, 28), cv2.FONT_HERSHEY_SIMPLEX, 0.72, (255, 255, 255), 2)
            cv2.putText(show, "SPACE=capture  S=skip  Q=quit", (12, 55), cv2.FONT_HERSHEY_SIMPLEX, 0.55, (180, 220, 180), 1)
            cv2.imshow("RobustAuth Admin Capture", show)
            key = cv2.waitKeyEx(30)
            key = key & 0xFF if key != -1 else -1
            if key in (ord("q"), ord("Q")):
                cap.release(); cv2.destroyAllWindows()
                return saved_total > 0
            if key in (ord("s"), ord("S")):
                break
            if key in (32, 13) and len(faces) > 0:
                out = student_folder / f"{angle}_{saved + 1}.jpg"
                cv2.imwrite(str(out), frame)
                saved += 1
                saved_total += 1
                print(f"Saved {out.name}")
    cap.release()
    cv2.destroyAllWindows()
    return saved_total > 0


def add_student():
    header("ADD STUDENT")
    name = input("Student full name: ").strip()
    if not name:
        print("Name cannot be empty."); pause(); return
    folder = folder_name(name)
    path = DATASET_DIR / folder
    if path.exists():
        ans = input(f"{name} already exists. Re-capture photos? (yes/no): ").strip().lower()
        if ans != "yes":
            return
        shutil.rmtree(path)
    path.mkdir(parents=True, exist_ok=True)
    print("Camera will capture center/left/right/up/down photos.")
    input("Press Enter to open camera...")
    if capture_photos(path, name):
        print("Photos saved. Now retrain model from option 2.")
    else:
        print("No photos captured.")
        if path.exists() and not any(path.iterdir()):
            path.rmdir()
    pause()


def retrain():
    header("RETRAIN MODEL")
    print("Training HOG+LBP SVM + LBPH ensemble. This may take a few minutes for large datasets.\n")
    train_model.train()
    pause()


def export_excel():
    header("EXPORT ATTENDANCE TO EXCEL")
    if not EXCEL_OK:
        print("openpyxl is missing. Run: pip install openpyxl")
        pause(); return
    csv_files = sorted(ATTENDANCE_DIR.glob("attendance_*.csv"))
    if not csv_files:
        print("No attendance CSV files found."); pause(); return
    for idx, file in enumerate(csv_files, start=1):
        print(f"[{idx}] {file.name}")
    print("[A] All files")
    choice = input("Choice: ").strip().upper()
    if choice == "A":
        selected = csv_files
    elif choice.isdigit() and 1 <= int(choice) <= len(csv_files):
        selected = [csv_files[int(choice) - 1]]
    else:
        print("Invalid choice."); pause(); return

    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    for csv_path in selected:
        with open(csv_path, newline="", encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
        if not rows:
            continue
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = csv_path.stem.replace("attendance_", "")[:31]
        headers = list(rows[0].keys())
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title = ws.cell(1, 1, f"Attendance Report - {ws.title}")
        title.font = Font(bold=True, size=14, color="1F4E79")
        title.alignment = center
        for c, header in enumerate(headers, start=1):
            cell = ws.cell(2, c, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.alignment = center
            cell.border = border
            ws.column_dimensions[chr(64 + min(c, 26))].width = max(14, min(35, len(header) + 6))
        for r, row in enumerate(rows, start=3):
            for c, header in enumerate(headers, start=1):
                cell = ws.cell(r, c, row.get(header, ""))
                cell.alignment = center
                cell.border = border
                if row.get("Status") == "Present":
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
        xlsx_path = csv_path.with_suffix(".xlsx")
        wb.save(xlsx_path)
        print(f"Saved {xlsx_path.name}")
    pause()


def show_students():
    header("REGISTERED STUDENTS")
    students = list_students()
    print(f"Total students: {len(students)}\n")
    for idx, student in enumerate(students, start=1):
        photos = len([p for p in student.iterdir() if p.suffix.lower() in {'.jpg', '.jpeg', '.png'}])
        print(f"{idx:03d}. {display_name(student.name):<38} photos={photos}")
    pause()


def delete_student():
    header("DELETE STUDENT")
    students = list_students()
    if not students:
        print("No students found."); pause(); return
    for idx, student in enumerate(students, start=1):
        print(f"[{idx}] {display_name(student.name)}")
    choice = input("Number to delete: ").strip()
    if not choice.isdigit() or not (1 <= int(choice) <= len(students)):
        print("Invalid choice."); pause(); return
    target = students[int(choice) - 1]
    ans = input(f"Delete {display_name(target.name)}? Type YES: ").strip()
    if ans == "YES":
        shutil.rmtree(target)
        print("Deleted. Retrain the model.")
    else:
        print("Cancelled.")
    pause()


def main():
    header("LOGIN")
    if input("Admin password: ").strip() != ADMIN_PASSWORD:
        print("Wrong password.")
        return
    while True:
        header()
        print(f"Students: {len(list_students())}")
        print("\n[1] Add student by webcam")
        print("[2] Retrain model")
        print("[3] Export attendance Excel")
        print("[4] List students")
        print("[5] Delete student")
        print("[Q] Quit")
        ch = input("\nChoice: ").strip().upper()
        if ch == "1": add_student()
        elif ch == "2": retrain()
        elif ch == "3": export_excel()
        elif ch == "4": show_students()
        elif ch == "5": delete_student()
        elif ch == "Q": break
        else:
            print("Invalid choice."); pause()


if __name__ == "__main__":
    main()
